# encoding: utf-8
'''
#-------------------------------------------------------------------
#
#                   @Project Name : RealJudgeStatus
#
#                   @Programmer   : Koshou
#
#                   @Start Date   : 2022/5/26
#
#                   @Last Update  : 2022/6/16
#
#-------------------------------------------------------------------
'''



import os, openpyxl, time, json, ndjson, datetime, shutil, sys
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import pandas as pd
from pandas.plotting import register_matplotlib_converters
register_matplotlib_converters()



FIG_FOLDER = 'fig'
CSV_FOLDER = 'csv'
ERROR_FOLDER = 'error'
OUTPUT_PATH_LIST = [FIG_FOLDER, CSV_FOLDER, ERROR_FOLDER]

MOX_SENSOR_LIST = ['TGS_2600', 'TGS_2602', 'TGS_2603', 'TGS_2610', 'TGS_2611', 'TGS_2620', 'TGS_2612', 'FIS_5100', 'FIS_5300', 'FIS_Q106', 'FIS_3004', 'FIS_Q201', 'FIS_3B00', 'FIS_6100',]
MOX_SENSOR__DR_LIST = [f"{i}_dr"for i in MOX_SENSOR_LIST]
COLORMAP = ['#FFBE7D', '#4E79A7', '#F28E2B', '#A0CBE8', '#59A14F', '#8CD17D', '#B6992D', '#F1CE63', '#499894', '#86BCB6', '#E15759', '#FF9D9A', '#79706E', '#BAB0AC', '#D37295', '#FABFD2', '#B07AA1', '#D4A6C8', '#9D7660', '#D7B5A6']
STATUS_DATE_FLAG_MAP = {
    'gas_in':None,
    'reaction_stable':None,
    'recovery':None
}
DEBUG_MODE = False
def printf(content, mode):
    if DEBUG_MODE:
        if mode == 0:
            print(content)
        if mode == 1:
            print(content, '', end='')
    
class RealJudgeStatus:
    def __init__(self, parameters_dict, logger):
        self.sensor_counter_map = {
        'TGS_2600_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'TGS_2602_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'TGS_2603_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'TGS_2610_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'TGS_2611_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'TGS_2620_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'TGS_2612_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_5100_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_5300_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_Q106_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_3004_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_Q201_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_3B00_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}},
        'FIS_6100_dr':{'gas_in':{'counter':0, 'flag' : 0}, 'reaction_stable':{'counter':0, 'flag' : 0}, 'recovery':{'counter':0, 'flag' : 0}}
        }
        self.parameters_dict = parameters_dict
        self.logger = logger
        self.time_list = []
        self.baseline = []
        self.rjs_time_list = []
        self.vol_data_list = []
        self.rjs_vol_data_list = []
        self.res_data_list = []
        self.rjs_res_data_list = []
        self.rjs_baseline_df = pd.DataFrame()
        self.rjs_rmdr_df = pd.DataFrame() # res max difference rate
        self.rjs_sadr_df = pd.DataFrame(columns= MOX_SENSOR__DR_LIST) # use to save stable times
        self.rjs_rsdr_df = pd.DataFrame() # res stable difference rate
        self.rjs_rms_df = pd.DataFrame() # res max slope
        self.rjs_r5dr_df = pd.DataFrame() # res 5s difference rate
        self.rjs_r10dr_df = pd.DataFrame() # res 10s difference rate
        self.rjs_r5s_df = pd.DataFrame() # res 5s slope
        self.rjs_r10s_df = pd.DataFrame() # res 10s slope
        self.rmdr_flag = False # res max difference rate
        self.rms_flag = False # res max slope
        self.feature_done = False # use for check feature done
        self.stable_status_flag = False # use for stable status or not stable
        self.create_baseline_info_flag = False # if called func turn true
        self.humi_list = []
        self.temp_list = []
        self.flow_list = []
        self.first_find_stable = False
        self.flag_flag = 0 # I only append the point of status start 
        self.append_flag = False # use for return once
        self.status_date_flag_map= {
            'esrtsd':{'gas_in':None, 'reaction_stable':None, 'recovery':None},
            'rjs':{'gas_in':None, 'reaction_stable':None, 'recovery':None}
        }
        self.feature_map = {
            'resistance_max_difference_rate' : {'df':self.rjs_rmdr_df, 'sort_name':'rmdr'},
            'resistance_stable_difference_rate' : {'df':self.rjs_rsdr_df, 'sort_name':'rsdr'},
            'resistance_max_slope' : {'df':self.rjs_rms_df, 'sort_name':'rms'},
            'resistance_5s_difference_rate' : {'df':self.rjs_r5dr_df, 'sort_name':'r5dr'},
            'resistance_10s_difference_rate' : {'df':self.rjs_r10dr_df, 'sort_name':'r10dr'},
            'resistance_5s_slope' : {'df':self.rjs_r5s_df, 'sort_name':'r5s'},
            'resistance_10s_slope' : {'df':self.rjs_r10s_df, 'sort_name':'r10s'}
        }

    def append_nessacry_data(self, each_line:dict):
        '''
        append which kind of data must to append at status1 to status7
        '''

        self.time_list.append(datetime.datetime.strptime(each_line['l_date'].split('.')[0], "%Y-%m-%dT%H:%M:%S"))
        self.vol_data_list.append(each_line['channel'][:14])
        self.humi_list.append(each_line['channel'][14])
        self.temp_list.append(each_line['channel'][15])
        self.flow_list.append(each_line['channel'][16])

    def append_rjs_data(self, each_line:dict):
        '''
        append data which use by real judge status
        '''
        
        self.rjs_time_list.append(datetime.datetime.strptime(each_line['l_date'].split('.')[0], "%Y-%m-%dT%H:%M:%S"))
        self.rjs_vol_data_list.append(each_line['channel'][:14])

        
    def append_flag_point(self, each_line:dict, status:str):
        '''
        append flag point
        '''
        if self.flag_flag == 0:
            self.status_date_flag_map['esrtsd'][status] = datetime.datetime.strptime(each_line['l_date'].split('.')[0], "%Y-%m-%dT%H:%M:%S")
            self.flag_flag = 1
      
    def vol_to_res_down(self, vol:float, rbias:float)->(float):
        '''
        calculate res which sensor location down side
        '''
        
        rsensor = vol*rbias/(5-vol)
        return rsensor

    def vol_to_res_up(self, vol:float, rbias:float)->(float):
        '''
        calculate res which sensor location up side
        '''
        
        rsensor = (5-vol)*rbias/vol
        return rsensor

    def transfer_to_resistance(self, voltage_list :list) -> (list):
        '''
        voltage data transfer to resistance
        '''
        
        res_list = []
        for k in range(len(voltage_list)):
            tmp_list = []
            for i in range(14):
                if i in range(7, 11, 1):
                    tmp_list.append(self.vol_to_res_down(voltage_list[k][i], self.third_presetvoltage_list[i]))
                else:
                    tmp_list.append(self.vol_to_res_up(voltage_list[k][i], self.third_presetvoltage_list[i]))
            res_list.append(tmp_list)
        return res_list
    
    def plot(self):
        '''
        plot two axis fig in each png file, one is raw status, anthoer is RealJudgeStatus 
        '''
        

        plt.figure(figsize=(20, 20))
        res_data_ary = np.asarray(self.res_data_list)
        # plot first fig
        ax1 = plt.subplot(2, 1, 1)
        for i in range(14):
            ax1.plot(self.time_list, res_data_ary[:, i], label=MOX_SENSOR_LIST[i], color=COLORMAP[i])
            
        for i in list(self.status_date_flag_map['esrtsd'].keys()):
            if self.status_date_flag_map['esrtsd'][i] == None:
                pass
            else:      
                ax1.axvline(x = self.status_date_flag_map['esrtsd'][i], ymin= 0, ymax= np.max(res_data_ary))
                text_content = "{:d}:{:02d}:{:02d}".format(self.status_date_flag_map['esrtsd'][i].hour, self.status_date_flag_map['esrtsd'][i].minute, self.status_date_flag_map['esrtsd'][i].second)
                plt.text(self.status_date_flag_map['esrtsd'][i], np.max(res_data_ary), text_content, fontsize=10, rotation = 30)
                ax1.legend(loc='upper left', bbox_to_anchor=(1, 0.5))
       
        ax1.grid()
        ax1.title.set_text('ESRTSD\n\n\n')

        # plot second fig
        ax2 = plt.subplot(2, 1, 2)
        for i in range(14):
            ax2.plot(self.time_list, res_data_ary[:, i], label=MOX_SENSOR_LIST[i], color=COLORMAP[i])
        
        for i in list(self.status_date_flag_map['rjs'].keys()):
            if self.status_date_flag_map['rjs'][i] == None:
                pass
            else:      
                ax2.axvline(x = self.status_date_flag_map['rjs'][i], ymin= 0, ymax= np.max(res_data_ary))
                text_content = "{:d}:{:02d}:{:02d}".format(self.status_date_flag_map['rjs'][i].hour, self.status_date_flag_map['rjs'][i].minute, self.status_date_flag_map['rjs'][i].second)
                plt.text(self.status_date_flag_map['rjs'][i], np.max(res_data_ary), text_content, fontsize=10, rotation = 30)
        
        ax2.legend(loc='upper left', bbox_to_anchor=(1, 0.5))
        ax2.grid()
        ax2.title.set_text('RealJudgeStatus\n\n\n')
        
        # plt.gcf().autofmt_xdate()
        tmplist = self.filename.split('\\')[1:]
        output_path = FIG_FOLDER
        for i in tmplist:
            output_path = os.path.join(output_path, i)
        plt.savefig(os.path.join(output_path.replace('ndjson', 'png')), dpi=100)
        # plt.show()  
        # plt.clf()
        plt.close()
        

    def split_status(self, i:dict):
        '''
        split status 1 to status 7
        '''
        
        # status = 0, initial
        # status = 1, on_first_preset_voltage
        # status = 2, on_second_preset_voltage
        # status = 3, on_round_start
        # status = 4, on_baseline_stable
        # status = 5, on_collecting
        # status = 6, on_stable
        # status = 7, on_recovery
        if i['name'] == 'control_status':
            if i['status_flag'] == 'on_first_preset_voltage':
                self.status = 1
            elif i['status_flag'] == 'on_second_preset_voltage':
                self.status = 2
            elif i['status_flag'] == 'on_round_start':
                self.status = 3
            elif i['status_flag'] == 'on_baseline_stable':
                self.status = 4
            elif i['status_flag'] == 'on_collecting':
                self.status = 5
                self.flag_flag = 0
            elif i['status_flag'] == 'on_stable':
                self.status = 6
                self.flag_flag = 0
            elif i['status_flag'] == 'on_recovery':
                self.status = 7
                self.flag_flag = 0
        if self.status == 1:
            if i['name'] == 'preset_meta':
                pass
            elif i['name'] == 'data':
                self.append_nessacry_data(i)
        if self.status == 2:
            if i['name'] == 'preset_meta':
                pass
            elif i['name'] == 'data':
                self.append_nessacry_data(i)
        if self.status == 3:
            if i['name'] == 'round_meta':
                self.label = i['label']
                self.analysis_term = i['global_addfield']['analysis_term']['value']
                self.provider = i['global_addfield']['provider']['value']
                if 'simple_id' in i['global_addfield']:
                    self.simple_id = i['global_addfield']['simple_id']['value']
                else:
                    self.simple_id = None
            elif i['name'] == 'preset_meta':
                self.third_presetvoltage_list = i['device_cfg']['channel_biasing_resistance']
            elif i['name'] == 'data':
                self.append_nessacry_data(i)
                
        if self.status == 4:
            if i['name'] == 'data':
                self.append_nessacry_data(i)
                self.append_rjs_data(i)
                self.baseline.append(i['channel'][:14])
        if self.status == 5:
            if i['name'] == 'data':
                self.append_nessacry_data(i)
                self.append_flag_point(i,'gas_in')
                self.append_rjs_data(i)
                if not self.create_baseline_info_flag:
                    res_baseline = self.transfer_to_resistance(self.baseline)
                    res_baseline_df = pd.DataFrame(res_baseline)
                    self.create_baseline_info(res_baseline_df)
                    # print(self.baseline_info_df)
                    
        if self.status == 6:
            if i['name'] == 'data':
                self.append_nessacry_data(i)
                self.append_flag_point(i, 'reaction_stable')
                self.append_rjs_data(i)
        if self.status == 7:
            if i['name'] == 'data':
                self.append_nessacry_data(i)
                self.append_flag_point(i, 'recovery')
                self.append_rjs_data(i)

    def check_modified(self, all_data:list):
        '''
        check file if modified
        '''
        
        self.status = 0
        skip_index = None
        modified_flag = False  
        for index, i in enumerate(all_data):
            if index == 0:
                if i['name'] == 'modify_header':
                    modified_flag = True
                    skip_index = i['skip_to']
                else:
                    modified_flag = False
                    
            if modified_flag:
                if index >= skip_index:  
                    self.split_status(i) 
                else:
                    pass
            else:   
                self.split_status(i)

    def test_dataload(self,filename:str):
        
        self.filename = filename
        with open(filename, 'r', encoding="utf-8") as f:
            print(filename)
            all_data = ndjson.load(f)            
            self.check_modified(all_data)
        self.res_data_list = self.transfer_to_resistance(self.vol_data_list)
        self.rjs_entrance()
        self.plot()
        
    def copy_error_file(self):
        '''
        copy error file
        '''
        
        tmplist = self.filename.split('\\')[1:]
        output_path = ERROR_FOLDER
        for i in tmplist:
            output_path = os.path.join(output_path, i)
        shutil.copyfile(self.filename, os.path.join(output_path))

    def dataload(self,filename:str, statistic_df:pd.DataFrame, features_df:pd.DataFrame) -> (pd.DataFrame):
        self.filename = filename
        try:
            with open(filename, 'r', encoding="utf-8") as f:
                print(filename)
                self.logger.logging(filename)
                all_data = ndjson.load(f)            
                self.check_modified(all_data)
            self.res_data_list = self.transfer_to_resistance(self.vol_data_list)
            features_df = self.rjs_entrance(features_df)
            self.plot()
            statistic_df = self.statistic_total(statistic_df)
            return statistic_df, features_df
        except Exception as e: 
            print(e)
            self.logger.logging('Exception')
            self.logger.logging(str(e))
            self.copy_error_file()
            return statistic_df, features_df
        
    def rolling_calculate(self, res_data_df:pd.DataFrame) -> (pd.DataFrame):
        '''
        calculate each two sec data, (St2 - St1))/St1
        '''
        
        difference_rate_df = pd.DataFrame(np.nan, index = np.arange(res_data_df.shape[0]), columns = MOX_SENSOR__DR_LIST)
        for cols, cols_dr in zip(MOX_SENSOR_LIST, MOX_SENSOR__DR_LIST):
            for index in range(res_data_df.shape[0]):
                if index == res_data_df.shape[0]-1:
                    break
                else:
                    difference_rate_df.at[index, cols_dr]  = (res_data_df.at[index+1, cols] - res_data_df.at[index, cols])/res_data_df.at[index, cols]
        # difference_rate_df.to_csv('dr.csv')
        return difference_rate_df

    def judgment_status(self, difference_rate_df:pd.DataFrame, res_data_df:pd.DataFrame, features_df:pd.DataFrame) -> (pd.DataFrame):
        '''
        rjs judgment status
        '''
        rjs_status = 4
        wait_stable_counter = 0 # use for count if not stable need to count to limit and recovery
        counter = 0
        for index in range(difference_rate_df.shape[0]):
            printf(index,1)
            sucess_gas_in_flag_num = 0
            sucess_reaction_stable_flag_num = 0
            sucess_recovery_flag_num = 0
            for cols_dr in MOX_SENSOR__DR_LIST:
                #--- judge gas in start ---#
                if rjs_status == 4:
                    if abs(float(difference_rate_df[index:index+1][cols_dr])) >= (self.parameters_dict[cols_dr]['gas_in']['threshold']*self.baseline_info_df.at[3, cols_dr]):
                        self.sensor_counter_map[cols_dr]['gas_in']['counter']+=1
                    else:
                        self.sensor_counter_map[cols_dr]['gas_in']['counter'] = 0
                        self.sensor_counter_map[cols_dr]['gas_in']['flag'] = 0
                    if self.sensor_counter_map[cols_dr]['gas_in']['counter'] >= self.parameters_dict[cols_dr]['gas_in']['conti_times']:
                        self.sensor_counter_map[cols_dr]['gas_in']['flag'] = 1
                        printf(self.sensor_counter_map[cols_dr]['gas_in']['counter'],0)
                #--- judge reaction stable start ---#
                if rjs_status == 5:
                    if abs(float(difference_rate_df[index:index+1][cols_dr])) <= (self.parameters_dict[cols_dr]['reaction_stable']['threshold']*self.baseline_info_df.at[3, cols_dr]):
                        self.sensor_counter_map[cols_dr]['reaction_stable']['counter']+=1
                    else:
                        self.sensor_counter_map[cols_dr]['reaction_stable']['counter'] = 0
                        self.sensor_counter_map[cols_dr]['reaction_stable']['flag'] = 0
                        self.clean_stable_value_each_sensor(index, cols_dr)
                    if self.sensor_counter_map[cols_dr]['reaction_stable']['counter'] >= (self.parameters_dict[cols_dr]['reaction_stable']['conti_times']*self.baseline_info_df.at[3, cols_dr]):
                        self.sensor_counter_map[cols_dr]['reaction_stable']['flag'] = 1
                        self.append_stable_value_each_sensor(index, cols_dr)
                #--- judge recovery start ---#
                if rjs_status == 6:
                    if abs(float(difference_rate_df[index:index+1][cols_dr])) >= self.parameters_dict[cols_dr]['recovery']['threshold']:
                        self.sensor_counter_map[cols_dr]['recovery']['counter']+=1
                    else:
                        self.sensor_counter_map[cols_dr]['recovery']['counter'] = 0
                        self.sensor_counter_map[cols_dr]['recovery']['flag'] = 0
                    if self.sensor_counter_map[cols_dr]['recovery']['counter'] >= self.parameters_dict[cols_dr]['recovery']['conti_times']:
                        self.sensor_counter_map[cols_dr]['recovery']['flag'] = 1
                        
                    if abs(float(difference_rate_df[index:index+1][cols_dr])) <= self.parameters_dict[cols_dr]['reaction_stable']['threshold']:
                        self.sensor_counter_map[cols_dr]['reaction_stable']['counter']+=1
                    else:
                        self.sensor_counter_map[cols_dr]['reaction_stable']['counter'] = 0
                        self.sensor_counter_map[cols_dr]['reaction_stable']['flag'] = 0
                        self.clean_stable_value_each_sensor(index, cols_dr)
                    if self.sensor_counter_map[cols_dr]['reaction_stable']['counter'] >= self.parameters_dict[cols_dr]['reaction_stable']['conti_times']:
                        self.sensor_counter_map[cols_dr]['reaction_stable']['flag'] = 1
                        self.append_stable_value_each_sensor(index, cols_dr)

                
                sucess_gas_in_flag_num += self.sensor_counter_map[cols_dr]['gas_in']['flag']
                sucess_reaction_stable_flag_num += self.sensor_counter_map[cols_dr]['reaction_stable']['flag']
                sucess_recovery_flag_num += self.sensor_counter_map[cols_dr]['recovery']['flag']
                # printf(self.sensor_counter_map[cols_dr]['reaction_stable']['flag'],1)
            printf('',0)
            
            if rjs_status == 4:
                printf(sucess_gas_in_flag_num,1)
                self.append_baseline_res_value(index, res_data_df)
                if sucess_gas_in_flag_num >= self.parameters_dict['judgment_sensor_num']:
                    # print(index, 'gas in here')
                    self.status_date_flag_map['rjs']['gas_in'] = datetime.datetime.strptime(str(difference_rate_df[index:index+1]['time'].values).split('.')[0].replace('[\'', ''), "%Y-%m-%dT%H:%M:%S")
                    rjs_status = 5
                    printf('gas in',0)

            elif rjs_status == 5:
                wait_stable_counter +=1
                if sucess_reaction_stable_flag_num >= self.parameters_dict['judgment_sensor_num']:
                    # print(index, 'reaction stable here')
                    self.status_date_flag_map['rjs']['reaction_stable'] = datetime.datetime.strptime(str(difference_rate_df[index:index+1]['time'].values).split('.')[0].replace('[\'', ''), "%Y-%m-%dT%H:%M:%S")
                    rjs_status = 6
                    self.stable_status_flag = True
                    printf('reaction stable',0)
                elif wait_stable_counter == self.parameters_dict['wait_stable_limit']:   
                    self.status_date_flag_map['rjs']['recovery'] = datetime.datetime.strptime(str(difference_rate_df[index:index+1]['time'].values).split('.')[0].replace('[\'', ''), "%Y-%m-%dT%H:%M:%S")
                    printf('1recovery',0)
                    rjs_status = 7
                    self.recovery = index
                
            elif rjs_status == 6:
                wait_stable_counter += 1
                if sucess_recovery_flag_num >= self.parameters_dict['judgment_sensor_num']:
                    # print(index, 'recovery here')
                    printf('2recovery',0)
                    self.status_date_flag_map['rjs']['recovery'] = datetime.datetime.strptime(str(difference_rate_df[index:index+1]['time'].values).split('.')[0].replace('[\'', ''), "%Y-%m-%dT%H:%M:%S")        
                    rjs_status = 7
                    self.recovery = index
                elif wait_stable_counter == self.parameters_dict['wait_stable_limit']:   
                    self.status_date_flag_map['rjs']['recovery'] = datetime.datetime.strptime(str(difference_rate_df[index:index+1]['time'].values).split('.')[0].replace('[\'', ''), "%Y-%m-%dT%H:%M:%S")
                    printf('1recovery',0)
                    rjs_status = 7
                    self.recovery = index
            # print(rjs_status, wait_stable_counter)       
            features_df = self.feature_extraction(index, rjs_status, res_data_df, wait_stable_counter, features_df)
        return features_df
        
    def append_stable_value_each_sensor(self, index:int, cols_dr:str):
        '''
        append res of each sensor judge at stable status
        '''
        self.rjs_sadr_df.at[index, cols_dr] =  1
        if not self.first_find_stable:
            self.gas_in_flag = index
            self.first_find_stable = True
            
    def clean_stable_value_each_sensor(self, index:int, cols_dr:str):
        '''
        if not stable, clean
        '''
        
        # self.rjs_sadr_df[cols_dr] =  0
        
        
    def append_baseline_res_value(self, index:int, res_data_df:pd.DataFrame):
        '''
        append all value at status 4(baseline stable)
        '''
        
        self.rjs_baseline_df = self.rjs_baseline_df.append(res_data_df[index:index+1])
        
    def baseline_avg(self) -> (pd.DataFrame):
        '''
        return average of value
        '''
        return self.rjs_baseline_df.mean().to_frame().T

    def resistance_max_difference_rate(self, index:int, res_data_df:pd.DataFrame):
        '''
        reaction max difference rate from gasin start, (Rt1-Rbaseline)/Rbaseline
        '''

        if not self.rmdr_flag:
            for sensor in MOX_SENSOR_LIST:
                self.rjs_rmdr_df.at[0, sensor] = (res_data_df.at[index,sensor] - self.baseline_avg_df.at[0, sensor])/self.baseline_avg_df.at[0, sensor]
            self.rmdr_flag = True
        else:
            for sensor in MOX_SENSOR_LIST:
                if (abs((res_data_df.at[index,sensor] - self.baseline_avg_df.at[0, sensor])/self.baseline_avg_df.at[0, sensor]) > abs(self.rjs_rmdr_df.at[0, sensor])):
                    self.rjs_rmdr_df.at[0, sensor] = (res_data_df.at[index,sensor] - self.baseline_avg_df.at[0, sensor])/self.baseline_avg_df.at[0, sensor]
        
    def resistance_stable_difference_rate(self, index:int, res_data_df:pd.DataFrame):
        '''
        reaction stable difference rate, each sensor get theirself stable value
        if this sensor no stable or stable contiunte time not enough, grab before recovery a period of times 
        if this sensor has two times stable status, choose the last status
        if rjs didnt judge statble status, replace with resistance max difference rate
        '''
        
        if self.stable_status_flag:
            self.rjs_sadr_df = self.rjs_sadr_df.fillna(0)
            for cols, raw_cols in zip(self.rjs_sadr_df.columns, MOX_SENSOR_LIST):
                alist = list(self.rjs_sadr_df[cols])
                first_find0_after_find1_flag = False
                first_find1_flag = False
                final_find_flag = False
                stable_conti_counter = 0
                index_saver = []
                for i in range(len(alist)-1, -1, -1):
                    if alist[i] == 1 and not first_find0_after_find1_flag:
                        stable_conti_counter += 1
                        index_saver.append(i)
                        first_find1_flag = True
                    if alist[i] == 0 and first_find1_flag and not first_find0_after_find1_flag:
                        first_find0_after_find1_flag = True
                        if stable_conti_counter < self.parameters_dict['each_sensor_statble_conti_times']:
                            # print('not enough stable')
                            first_find0_after_find1_flag = False
                            first_find1_flag = False
                            index_saver = []
                        else:
                            final_find_flag = True
                            # print('find stable')
                        stable_conti_counter = 0
                    # print(i, alist[i], stable_conti_counter)
                if final_find_flag:
                    # print(index_saver)
                    new_index_saver = [i+self.gas_in_flag for i in index_saver]
                    # print(res_data_df[raw_cols][new_index_saver[-1]:new_index_saver[0]+1].mean())
                    self.rjs_rsdr_df.at[0, cols] =(res_data_df[raw_cols][new_index_saver[-1]:new_index_saver[0]+1].mean()- self.baseline_avg_df.at[0, raw_cols])/self.baseline_avg_df.at[0, raw_cols]
                else:
                    # print(res_data_df[raw_cols][self.recovery-self.parameters_dict['not_stable_before_recovery_num']:self.recovery].mean())
                    self.rjs_rsdr_df.at[0, cols] = (res_data_df[raw_cols][self.recovery-self.parameters_dict['not_stable_before_recovery_num']:self.recovery].mean()- self.baseline_avg_df.at[0, raw_cols])/self.baseline_avg_df.at[0, raw_cols]
        else:
            self.rjs_rsdr_df = self.rjs_rmdr_df.copy()
            self.feature_map['resistance_stable_difference_rate']['df'] = self.rjs_rmdr_df.copy()



        
    def resistance_max_slope(self, index:int, res_data_df:pd.DataFrame):
        '''
        reaction max slope from gasin start, (Rt2-Rt1)/1s
        '''
        
        if not self.rms_flag:
            for sensor in MOX_SENSOR_LIST:
                self.rjs_rms_df.at[0, sensor] = (res_data_df.at[index,sensor] - res_data_df.at[index-1,sensor])/1
            self.rms_flag = True
        else:
            for sensor in MOX_SENSOR_LIST:
                if (abs((res_data_df.at[index,sensor] - res_data_df.at[index-1,sensor])/1) > abs(self.rjs_rms_df.at[0, sensor])):
                    self.rjs_rms_df.at[0, sensor] = (res_data_df.at[index,sensor] - res_data_df.at[index-1,sensor])/1
               
    def resistance_5s_difference_rate(self, index:int, res_data_df:pd.DataFrame, counter:int):
        '''
        reaction difference rate from gasin start at 5s, (Rt5-Rbaseline)/Rbaseline
        '''
        
        if counter == 4:
            for sensor in MOX_SENSOR_LIST:
                self.rjs_r5dr_df.at[0, sensor] = (res_data_df.at[index,sensor] - self.baseline_avg_df.at[0, sensor])/self.baseline_avg_df.at[0, sensor]
    
    def resistance_10s_difference_rate(self, index:int, res_data_df:pd.DataFrame, counter:int):
        '''
        reaction difference rate from gasin start at 10s, (Rt10-Rbaseline)/Rbaseline
        '''
        
        if counter == 9:
            for sensor in MOX_SENSOR_LIST:
                self.rjs_r10dr_df.at[0, sensor] = (res_data_df.at[index,sensor] - self.baseline_avg_df.at[0, sensor])/self.baseline_avg_df.at[0, sensor]
      
    def resistance_5s_slope(self, index:int, res_data_df:pd.DataFrame, counter:int):
        '''
        reaction slope from gasin start at 5s, (Rt5-Rt4)/1s
        '''
        if counter == 4:
            for sensor in MOX_SENSOR_LIST:
                self.rjs_r5s_df.at[0, sensor] = (res_data_df.at[index,sensor] - res_data_df.at[index-1,sensor])/1
                
    def resistance_10s_slope(self, index:int, res_data_df:pd.DataFrame, counter:int):
        '''
        reaction slope from gasin start at 10s, (Rt10-Rt9)/1s
        '''
        if counter == 9:
            for sensor in MOX_SENSOR_LIST:
                self.rjs_r10s_df.at[0, sensor] = (res_data_df.at[index,sensor] - res_data_df.at[index-1,sensor])/1

    def columns_rename(self):
        '''
        rename cols name of each feature
        '''

        for key in self.feature_map.keys():
            sortname = self.feature_map[key]['sort_name']
            re_cols = [f"{i}_{sortname}" for i in MOX_SENSOR_LIST]
            self.feature_map[key]['df'].columns = re_cols

    
    
    def feature_extraction(self, index:int, rjs_status:int, res_data_df:pd.DataFrame, counter:int, features_df:pd.DataFrame) -> (pd.DataFrame):
        '''
        feature extraction and then set output dataframe
        '''
        
        total_df = pd.DataFrame()
        if rjs_status == 4:
            self.baseline_avg_df = self.baseline_avg()
        if rjs_status == 5 or rjs_status == 6:
            self.resistance_max_difference_rate(index, res_data_df) 
            # self.resistance_stable_difference_rate(index, res_data_df) 
            self.resistance_max_slope(index, res_data_df) 
            self.resistance_5s_difference_rate(index, res_data_df, counter)  
            self.resistance_10s_difference_rate(index, res_data_df, counter)  
            self.resistance_5s_slope(index, res_data_df, counter)  
            self.resistance_10s_slope(index, res_data_df, counter) 
        if rjs_status == 7:
            if not self.feature_done:
                self.resistance_stable_difference_rate(index, res_data_df) 
                self.feature_done = True
            else:
                total_features_df_list = []
                self.columns_rename()
                for key in self.feature_map.keys():
                    total_features_df_list.append(self.feature_map[key]['df'])
                total_df = pd.concat(total_features_df_list, axis = 1)
                if self.append_flag:
                    return features_df
                else:
                    total_df.insert(0, 'file', self.filename)
                    total_df.insert(1, 'label', self.label)
                    features_df = features_df.append(total_df)
                    self.append_flag = True   
                    return features_df
        return features_df

        
        
    def save_threshold_csv(self, difference_rate_df:pd.DataFrame):
        '''
        save threshold of each time to excel
        '''
        tmplist = self.filename.split('\\')[1:]
        output_path = CSV_FOLDER
        for i in tmplist:
            output_path = os.path.join(output_path, i)
        excel_path = os.path.join(output_path.replace('.ndjson', '.xlsx'))
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active
        rows = dataframe_to_rows(difference_rate_df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws1.cell(row=r_idx, column=c_idx, value=value)
        
        ws2 = wb1.create_sheet(title='baseline_info') 
        rows = dataframe_to_rows(self.baseline_info_df)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        
        wb1.save(excel_path)

    def create_esrtsd_timeflag(self, rjs_time_df:pd.DataFrame) -> pd.DataFrame:
        rjs_time_df.insert(0, 'esrtsd_control_status', np.NaN)
        rjs_time_df['esrtsd_control_status'] = rjs_time_df['esrtsd_control_status'].astype(str)
        for index in range(len(rjs_time_df)):
            for i in self.status_date_flag_map['esrtsd'].keys():
                if self.status_date_flag_map['esrtsd'][i] == rjs_time_df.at[index, 'time']:
                    # print(self.status_date_flag_map['esrtsd'][i])
                    rjs_time_df.at[index,'esrtsd_control_status'] = i
        return rjs_time_df
    
    def create_baseline_info(self, df):
        '''
        including max, min, avg and (max-min)/avg
        this func need to called at gas in status, because baseline stable was iteration
        '''
        
        baseline_max = df.max().to_numpy()
        baseline_min = df.min().to_numpy()
        baseline_avg = df.mean().to_numpy()
        baseline_factor = (baseline_max - baseline_min)/baseline_avg

        total_ary = [baseline_max, baseline_min, baseline_avg, baseline_factor]
        baseline_info_df = pd.DataFrame(total_ary, columns= MOX_SENSOR__DR_LIST)
        description = ["baseline_max", "baseline_min", "baseline_avg", "baseline_factor"]
        baseline_info_df.insert(0, 'description', description)
        self.baseline_info_df = baseline_info_df
        self.create_baseline_info_flag = True

    
    def rjs_entrance(self, features_df:pd.DataFrame) -> (pd.DataFrame):
        '''
        the entrance of real judge status
        '''
        self.rjs_res_data_list = self.transfer_to_resistance(self.rjs_vol_data_list)
        res_data_df = pd.DataFrame(self.rjs_res_data_list, columns= MOX_SENSOR_LIST) #res_data_df no time
        rjs_time_df = pd.DataFrame(self.rjs_time_list, columns=['time'])
        rjs_time_df =  self.create_esrtsd_timeflag(rjs_time_df)
        # print(rjs_time_df)
        res = self.rolling_calculate(res_data_df)
        difference_rate_df = pd.concat([rjs_time_df, res], axis = 1) # difference_rate_df have time
        self.save_threshold_csv(difference_rate_df)
        
        features_df = self.judgment_status(difference_rate_df, res_data_df, features_df)
        return features_df


    def statistic_total(self, statistic_df:pd.DataFrame):
        '''
        store esrtsd status time and rjs judge status time for later analysis 
        '''
        
        statistic_cols = list(['file']) + list(self.status_date_flag_map['rjs'].keys())
        statistic_esrtsd_value = list([self.filename]) + list(self.status_date_flag_map['esrtsd'].values())
        statistic_rjs_value = list([self.filename]) + list(self.status_date_flag_map['rjs'].values())
        tmp_esrtsd_df = pd.DataFrame([statistic_esrtsd_value], columns = statistic_cols)
        tmp_rjs_df = pd.DataFrame([statistic_rjs_value], columns = statistic_cols)
        statistic_df = statistic_df.append(tmp_esrtsd_df)
        statistic_df = statistic_df.append(tmp_rjs_df)
        return statistic_df
        
        
        
def copy_all_folder_path(input_folder:str):
    '''
    copy all folder path input to output
    '''
    
    for path in OUTPUT_PATH_LIST:
        for root, dirs, files in os.walk(input_folder):
            path_elements = root.split('\\')
            copy_root = root.split('\\')[0].replace(input_folder,path)
            for i in range(1,len(path_elements)):
                copy_root = os.path.join(copy_root, path_elements[i])
            if not os.path.exists(copy_root):
                os.mkdir(copy_root)
                

 



def load_parameters() -> (dict):
    with open('parameters.json') as json_file:
        data = json.load(json_file) 
    return data

def reset_col_length(filename):
    '''
    modify len of cols, colors, and transfer nan to blank
    '''
    
    wb=openpyxl.load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        dataframe = pd.read_excel(filename).fillna('-')
        dataframe.loc[len(dataframe)]=list(dataframe.columns)
        # modify len of cols
        for col in dataframe.columns:				
            index=list(dataframe.columns).index(col)
            letter=openpyxl.utils.get_column_letter(index+1)
            collen=dataframe[col].apply(lambda x:len(str(x).encode())).max()	
            ws.column_dimensions[letter].width=collen*1.2+4	
            # color
            if col in ['total_num', 'gas_in_sucess', 'recovery_sucess']:
                fille = openpyxl.styles.PatternFill('solid', fgColor='00FFFF00') 
                for i in range(1,4):
                    ws.cell(row = i, column = index+1).fill = fille
        # nan to blank
        for row in ws:
            for i in row:
                if i.value == 'nan':
                    i.value = ''
    wb.save(filename)
	
def calculate_statistic_result(statistic_df:pd.DataFrame, parameters) -> (pd.DataFrame):
    '''
    calculate statistic result
    '''
    
    unixt_df = pd.DataFrame()
    for status in list(STATUS_DATE_FLAG_MAP.keys()):
        for index, i in enumerate(pd.to_datetime(statistic_df[status])):
            if i is pd.NaT:
                unixt_df.at[index, status+'_unixt'] = pd.NaT
            else:
                unix_time = time.mktime(i.timetuple())
                unixt_df.at[index, status+'_unixt'] = unix_time
    gas_in_counter = 0
    recovery_counter = 0
    for index, i in enumerate(range(0, len(statistic_df), 2)):
        gas_in_difference = unixt_df.at[i+1,'gas_in_unixt'] - unixt_df.at[i,'gas_in_unixt']
        recovery_difference = unixt_df.at[i+1,'recovery_unixt'] - unixt_df.at[i,'recovery_unixt']
        if gas_in_difference is pd.NaT or recovery_difference is pd.NaT:
            pass
        else:
            if abs(gas_in_difference) <= parameters['statistic_threshold']['gas_in']:
                gas_in_counter += 1
            if abs(recovery_difference) <= parameters['statistic_threshold']['recovery']:
                recovery_counter += 1
    total_num = len(statistic_df)/2
    statistic_df.at[0,'total_num'] = total_num
    statistic_df.at[0,'gas_in_sucess'] = gas_in_counter
    statistic_df.at[0,'recovery_sucess'] = recovery_counter
    statistic_df = statistic_df.astype(str)
    statistic_df.at[1,'gas_in_sucess'] = str(round(gas_in_counter/total_num*100, 2))+'%'
    statistic_df.at[1,'recovery_sucess'] = str(round(recovery_counter/total_num*100, 2)) +'%'
    return statistic_df
        
        



class Logger:
    
    def __init__(self, log_filename):
        self.log_path = 'log'
        self.log_filename = log_filename
        self.log_full_path = self.log_path +'\\'+ self.log_filename
        self.make_log_folder()
        
    
    def make_log_folder(self):            
        if not os.path.exists(os.path.join(os.getcwd(), self.log_path)):
            os.mkdir(os.path.join(os.getcwd(), self.log_path))

    
    def make_log_file(self):
        now = datetime.datetime.now()
        current_time = now.strftime("%H-%M-%S")
        with open(self.log_full_path , 'a') as f:
            f.write(current_time)
            f.write('log_start\n')
    
    
    
    def logging(self, message : str):
        '''
        log dump and open file write message in 
        '''
        
        now = datetime.datetime.now()
        current_time = now.strftime("%H-%M-%S")
        with open(self.log_full_path , 'a') as f:
            f.write(current_time)
            f.write(message)
            f.write('\n')
    
        pass
                


                





